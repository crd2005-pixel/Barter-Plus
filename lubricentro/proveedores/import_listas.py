# -*- coding: utf-8 -*-
"""
proveedores/import_listas.py
UI para importar Excel, ver listas e ítems y gestionar % por LISTA/MARCA.
Tras guardar %, se abre la lista con los % recién guardados "forzados".
Si se fuerza, se IGNORAN overrides por marca en esa apertura puntual.
En aperturas normales, siguen aplicando overrides por marca.
"""

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QDialog, QHeaderView, QFileDialog, QComboBox,
    QFormLayout, QLineEdit, QMessageBox, QTextEdit, QSpinBox, QDoubleSpinBox,
    QCheckBox, QGroupBox
)
from PyQt5.QtCore import Qt
    # noqa
from sqlalchemy import select, func, literal_column
from datetime import datetime
import openpyxl, os

# DB proveedores (listas/items/proveedores)
from .bootstrap import bootstrap
_ns = bootstrap()
Proveedor = _ns["Proveedor"]
ListaPrecioProveedor = _ns["ListaPrecioProveedor"]
ItemListaProveedor = _ns["ItemListaProveedor"]
SessionLocal = _ns["SessionLocal"]

# DB principal (si más adelante querés actualizar costo de productos)
from db import SessionLocal as AppSession, Producto

# Capa de % y cálculos
from .costos import (
    save_lista_percents, save_marca_percents, delete_marca_percents,
    get_lista_percents, get_marca_percents, calc_line
)

_PROV_OPTIONAL_FIELDS = [
    "telefono", "email", "direccion", "contacto",
    "notas", "observaciones", "cuit", "razon_social"
]

def _has_attr(model_cls, attr: str) -> bool:
    try:
        getattr(model_cls, attr); return True
    except Exception:
        return False

def _get_any(obj, *attrs):
    for a in attrs:
        try:
            return getattr(obj, a)
        except Exception:
            continue
    return None

# ---------------- Diálogo: Ítems de una lista ----------------
class _ItemsListaDialog(QDialog):
    """
    force_list_percents=(desc, iva) => usa exactamente esos % para TODA la lista
    e IGNORA overrides por marca en esta apertura puntual.
    force_list_percents=None => lee % de storage (lista y luego marca).
    """
    def __init__(self, lista_id: int, parent=None, force_list_percents=None):
        super().__init__(parent)
        self._lid = int(lista_id)
        self._forced = None
        if isinstance(force_list_percents, tuple) and len(force_list_percents) == 2:
            try:
                self._forced = (float(force_list_percents[0]), float(force_list_percents[1]))
            except Exception:
                self._forced = None

        self.setWindowTitle("Productos de la lista")
        self.resize(1040, 640)

        lay = QVBoxLayout(self)
        self.lbl = QLabel("", self); lay.addWidget(self.lbl)

        self.tbl = QTableWidget(self)
        self.tbl.setColumnCount(9)
        self.tbl.setHorizontalHeaderLabels(
            ["Código","Descripción","Marca","Presentación","Info Extra","Precio lista (costo)","Con desc.","+ IVA","Rubro"]
        )
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setSelectionMode(QTableWidget.SingleSelection)
        self.tbl.setSortingEnabled(True)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        lay.addWidget(self.tbl)

        self._load_header()
        self._load_items()

    def _base_percents(self):
        if self._forced is not None:
            # forzado por “Guardar %”
            d = max(0.0, min(99.99, self._forced[0]))
            v = max(0.0, min(99.99, self._forced[1]))
            return d, v
        # apertura normal: desde storage
        d, v = get_lista_percents(self._lid)
        return d, v

    def _load_header(self):
        with SessionLocal() as s:
            q = (select(ListaPrecioProveedor.id, ListaPrecioProveedor.nombre,
                        ListaPrecioProveedor.fecha_creacion, ListaPrecioProveedor.proveedor_id,
                        func.count(ItemListaProveedor.id))
                 .join(ItemListaProveedor, ItemListaProveedor.lista_id==ListaPrecioProveedor.id, isouter=True)
                 .where(ListaPrecioProveedor.id==self._lid).group_by(ListaPrecioProveedor.id))
            row = s.execute(q).first()
        d, v = self._base_percents()
        if row:
            _, nombre, fecha, pid, n = row
            with SessionLocal() as s:
                prov = s.get(Proveedor, pid) if pid else None
            pn = _get_any(prov,"nombre") if prov else ""
            ft = fecha.strftime("%Y-%m-%d") if hasattr(fecha,"strftime") and fecha else ""
            self.lbl.setText(f"Proveedor: {pn} | Lista: {nombre or ''} | Fecha: {ft} | Ítems: {n or 0} | % en uso -> Desc:{d:.2f}% IVA:{v:.2f}%")
        else:
            self.lbl.setText(f"Lista ID: {self._lid} | % en uso -> Desc:{d:.2f}% IVA:{v:.2f}%")

    def _load_items(self):
        code_col  = getattr(ItemListaProveedor, "producto_codigo", None) or getattr(ItemListaProveedor, "codigo")
        desc_col  = getattr(ItemListaProveedor, "descripcion")
        marca_col = getattr(ItemListaProveedor, "marca")
        rubro_col = getattr(ItemListaProveedor, "rubro_detectado", None) or literal_column("NULL")
        pres_col = getattr(ItemListaProveedor, "presentacion", None) or literal_column("NULL")
        extra_col = getattr(ItemListaProveedor, "info_extra", None) or literal_column("NULL")

        has_prec  = hasattr(ItemListaProveedor, "prec")

        with SessionLocal() as s:
            price_col = getattr(ItemListaProveedor, "prec") if has_prec else getattr(ItemListaProveedor, "precio")
            stmt = select(code_col, desc_col, marca_col, pres_col, extra_col, price_col.label("base"), rubro_col.label("rubro")) \
                   .where(ItemListaProveedor.lista_id == self._lid)
            rows = s.execute(stmt).all()

        d_base, v_base = self._base_percents()
        ignore_brand_overrides = (self._forced is not None)  # <<< CAMBIO CLAVE

        self.tbl.setRowCount(0)
        for codigo, desc, marca, pres, extra, base, rubro in rows:
            # Si venimos forzados, NO consultamos overrides por marca.
            if ignore_brand_overrides:
                d_eff, v_eff = d_base, v_base
            else:
                md, mv = get_marca_percents(self._lid, (marca or ""))
                d_eff = (md if md is not None else d_base)
                v_eff = (mv if mv is not None else v_base)

            basef, con_desc, mas_iva = calc_line(base, d_eff, v_eff)

            r = self.tbl.rowCount(); self.tbl.insertRow(r)
            self.tbl.setItem(r, 0, QTableWidgetItem("" if codigo is None else str(codigo)))
            self.tbl.setItem(r, 1, QTableWidgetItem("" if desc   is None else str(desc)))
            self.tbl.setItem(r, 2, QTableWidgetItem("" if marca  is None else str(marca)))
            self.tbl.setItem(r, 3, QTableWidgetItem("" if pres   is None else str(pres)))
            self.tbl.setItem(r, 4, QTableWidgetItem("" if extra  is None else str(extra)))
            self.tbl.setItem(r, 5, QTableWidgetItem(f"{basef:.2f}"))
            self.tbl.setItem(r, 6, QTableWidgetItem(f"{con_desc:.2f}"))
            self.tbl.setItem(r, 7, QTableWidgetItem(f"{mas_iva:.2f}"))
            self.tbl.setItem(r, 8, QTableWidgetItem("" if rubro  is None else str(rubro)))
        self.tbl.resizeColumnsToContents()

# ---------------- Asistente de mapeo Excel ----------------
class _ImportWizardDialog(QDialog):
    def __init__(self, xlsx_path: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Importar lista – Mapeo de columnas")
        self.resize(800, 560)
        self._wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        self._headers = []; self._rows = []
        lay = QVBoxLayout(self)

        fila1 = QHBoxLayout()
        fila1.addWidget(QLabel("Hoja:"))
        self.cbo_sheet = QComboBox(self); [self.cbo_sheet.addItem(n) for n in self._wb.sheetnames]
        fila1.addWidget(self.cbo_sheet); fila1.addStretch(1)
        lay.addLayout(fila1)

        grp = QGroupBox("Encabezado"); gl = QHBoxLayout(grp)
        gl.addWidget(QLabel("Fila de encabezado:"))
        self.sp_head = QSpinBox(self); self.sp_head.setRange(1, 9999); self.sp_head.setValue(1)
        gl.addWidget(self.sp_head); gl.addStretch(1)
        lay.addWidget(grp)

        mapg = QGroupBox("Mapeo de columnas"); fm = QFormLayout(mapg)
        self.cbo_cod = QComboBox(self); self.cbo_desc = QComboBox(self)
        self.cbo_marca = QComboBox(self); self.cbo_prec = QComboBox(self)
        self.cbo_pres = QComboBox(self); self.cbo_extra = QComboBox(self)
        fm.addRow("Código:", self.cbo_cod); fm.addRow("Descripción:", self.cbo_desc)
        fm.addRow("Marca:", self.cbo_marca); fm.addRow("Precio:", self.cbo_prec)
        fm.addRow("Presentación:", self.cbo_pres); fm.addRow("Info Extra:", self.cbo_extra)
        lay.addWidget(mapg)

        self.tbl = QTableWidget(self)
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        lay.addWidget(self.tbl)

        fila_b = QHBoxLayout()
        self.btn_ok = QPushButton("Importar"); self.btn_no = QPushButton("Cancelar")
        fila_b.addStretch(1); fila_b.addWidget(self.btn_ok); fila_b.addWidget(self.btn_no)
        lay.addLayout(fila_b)

        self.cbo_sheet.currentIndexChanged.connect(self._reload)
        self.sp_head.valueChanged.connect(self._reload)
        self.btn_ok.clicked.connect(self.accept); self.btn_no.clicked.connect(self.reject)
        self._reload()

    def _reload(self):
        ws = self._wb[self.cbo_sheet.currentText()]
        hrow = self.sp_head.value()
        rows = list(ws.iter_rows(values_only=True))
        heads, body = [], []
        for i, row in enumerate(rows, start=1):
            vals = [("" if v is None else str(v).strip()) for v in row]
            if i == hrow: heads = [v.lower() for v in vals]
            elif i > hrow and any(v != "" for v in vals): body.append(vals)
        self._headers, self._rows = heads, body

        def fill(cbo, prefer):
            cbo.clear()
            for i,h in enumerate(self._headers): cbo.addItem(f"{i}: {h}", i)
            pick = None
            for p in prefer:
                for i,h in enumerate(self._headers):
                    if p in h: pick = i; break
                if pick is not None: break
            if pick is None and cbo.count(): pick = 0
            if pick is not None: cbo.setCurrentIndex(pick)

        fill(self.cbo_cod,  ["codigo","código","sku","articulo","item"])
        fill(self.cbo_desc, ["descripcion","descripción","detalle","producto","nombre"])
        fill(self.cbo_marca,["marca","brand"])
        fill(self.cbo_prec, ["precio","pvp","neto","importe","valor","unitario"])
        fill(self.cbo_pres, ["presentacion","presentación","tamaño","medida","contenido"])
        fill(self.cbo_extra,["extra","info","observacion","nota"])

        self.tbl.clear(); self.tbl.setColumnCount(len(self._headers))
        self.tbl.setHorizontalHeaderLabels(self._headers)
        sample = self._rows[:100]; self.tbl.setRowCount(len(sample))
        for r,row in enumerate(sample):
            for c,v in enumerate(row):
                self.tbl.setItem(r,c,QTableWidgetItem("" if v is None else str(v)))
        self.tbl.resizeColumnsToContents()

    def selections(self):
        return {
            "sheet": self.cbo_sheet.currentText(),
            "header_row": self.sp_head.value(),
            "idx_codigo": self.cbo_cod.currentData(),
            "idx_desc":   self.cbo_desc.currentData(),
            "idx_marca":  self.cbo_marca.currentData(),
            "idx_precio": self.cbo_prec.currentData(),
            "idx_pres":   self.cbo_pres.currentData(),
            "idx_extra":  self.cbo_extra.currentData(),
        }

# ---------------- Pestaña principal ----------------
class ImportadorListasTab(QWidget):
    def __init__(self, on_after_import=None, parent=None):
        super().__init__(parent)
        self._on_after_import = on_after_import
        self._ui(); self._bind(); self._cargar_proveedores(); self._load_listas()

    def _ui(self):
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        self.cbo_prov = QComboBox(self)
        self.btn_new  = QPushButton("Nuevo")
        self.btn_edit = QPushButton("Editar")
        self.btn_view = QPushButton("Ver")
        self.btn_del_list = QPushButton("Eliminar lista")
        top.addWidget(QLabel("Proveedor:")); top.addWidget(self.cbo_prov, 2)
        top.addSpacing(8); top.addWidget(self.btn_new); top.addWidget(self.btn_edit); top.addWidget(self.btn_view)
        top.addStretch(1); top.addWidget(self.btn_del_list)
        lay.addLayout(top)

        box = QGroupBox("Porcentajes"); vbox = QVBoxLayout()

        row_l = QHBoxLayout()
        self.sp_desc_l = QDoubleSpinBox(self); self.sp_desc_l.setRange(0, 99.99); self.sp_desc_l.setDecimals(2); self.sp_desc_l.setSuffix(" %")
        self.sp_iva_l  = QDoubleSpinBox(self); self.sp_iva_l .setRange(0, 99.99); self.sp_iva_l .setDecimals(2); self.sp_iva_l .setSuffix(" %")
        self.btn_save_l = QPushButton("Guardar % en lista")
        row_l.addWidget(QLabel("Desc.% (lista):")); row_l.addWidget(self.sp_desc_l)
        row_l.addSpacing(10)
        row_l.addWidget(QLabel("IVA% (lista):"));  row_l.addWidget(self.sp_iva_l)
        row_l.addStretch(1); row_l.addWidget(self.btn_save_l)
        vbox.addLayout(row_l)

        row_m = QHBoxLayout()
        self.txt_marca = QLineEdit(self); self.txt_marca.setPlaceholderText("Marca exacta")
        self.sp_desc_m = QDoubleSpinBox(self); self.sp_desc_m.setRange(0, 99.99); self.sp_desc_m.setDecimals(2); self.sp_desc_m.setSuffix(" %")
        self.sp_iva_m  = QDoubleSpinBox(self); self.sp_iva_m .setRange(0, 99.99); self.sp_iva_m .setDecimals(2); self.sp_iva_m .setSuffix(" %")
        self.btn_save_m = QPushButton("Guardar % marca"); self.btn_del_m = QPushButton("Borrar % marca")
        row_m.addWidget(QLabel("Marca:")); row_m.addWidget(self.txt_marca, 2)
        row_m.addWidget(QLabel("Desc.%:")); row_m.addWidget(self.sp_desc_m)
        row_m.addWidget(QLabel("IVA%:"));  row_m.addWidget(self.sp_iva_m)
        row_m.addStretch(1); row_m.addWidget(self.btn_save_m); row_m.addWidget(self.btn_del_m)
        vbox.addLayout(row_m)

        box.setLayout(vbox); lay.addWidget(box)

        act = QHBoxLayout()
        self.btn_import = QPushButton("Importar Excel")
        self.btn_open_items = QPushButton("Abrir productos…")
        self.btn_reload = QPushButton("Actualizar listas")
        self.chk_apply_cost = QCheckBox("Al importar, actualizar costo con “+ IVA”")
        self.chk_apply_cost.setChecked(True)
        self.lbl_state = QLabel("")
        act.addWidget(self.btn_import); act.addWidget(self.btn_open_items); act.addWidget(self.btn_reload)
        act.addStretch(1); act.addWidget(self.chk_apply_cost); act.addSpacing(12); act.addWidget(self.lbl_state)
        lay.addLayout(act)

        self.tbl = QTableWidget(self)
        self.tbl.setColumnCount(7)
        self.tbl.setHorizontalHeaderLabels(["✓","ID","Proveedor","Nombre","Fecha","Ítems","Últ. actualización"])
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setSelectionMode(QTableWidget.SingleSelection)
        self.tbl.setSortingEnabled(True)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        lay.addWidget(self.tbl)

        self.setLayout(lay)

    def _bind(self):
        self.btn_reload.clicked.connect(self._load_listas)
        self.btn_import.clicked.connect(self._importar_excel)
        self.btn_open_items.clicked.connect(self._open_items)
        self.btn_new.clicked.connect(self._nuevo_proveedor)
        self.btn_edit.clicked.connect(self._editar_proveedor)
        self.btn_view.clicked.connect(self._ver_proveedor)
        self.btn_del_list.clicked.connect(self._delete_lista)
        self.btn_save_l.clicked.connect(self._on_save_lista)
        self.btn_save_m.clicked.connect(self._on_save_marca)
        self.btn_del_m.clicked.connect(self._on_del_marca)
        self.tbl.doubleClicked.connect(self._open_items)

    # ---------- helpers ----------
    def _lista_sel(self):
        r = self.tbl.currentRow()
        if r < 0: return None
        it = self.tbl.item(r, 1)
        if not it: return None
        val = it.data(Qt.UserRole) or it.text()
        try: return int(val)
        except Exception: return None

    def _lista_tildada(self):
        ids = []
        for r in range(self.tbl.rowCount()):
            it = self.tbl.item(r, 0)
            if it and it.checkState() == Qt.Checked:
                idi = self.tbl.item(r, 1)
                if idi:
                    v = idi.data(Qt.UserRole) or idi.text()
                    try: ids.append(int(v))
                    except Exception: pass
        return ids[0] if len(ids) == 1 else None

    def _clear_checks_except(self, keep_row):
        for r in range(self.tbl.rowCount()):
            if r == keep_row: continue
            it = self.tbl.item(r, 0)
            if it: it.setCheckState(Qt.Unchecked)

    # ---------- proveedores ----------
    def _cargar_proveedores(self):
        self.cbo_prov.clear()
        with SessionLocal() as s:
            rows = s.execute(select(Proveedor.id, Proveedor.nombre).order_by(Proveedor.nombre.asc())).all()
        for pid, name in rows:
            self.cbo_prov.addItem(name or "", pid)

    def _nuevo_proveedor(self):
        dlg = QDialog(self); dlg.setWindowTitle("Nuevo proveedor"); dlg.resize(520, 420)
        v = QVBoxLayout(dlg); form = QFormLayout()
        txt_nombre = QLineEdit(dlg); form.addRow("Nombre:", txt_nombre)
        inputs = {}
        for f in _PROV_OPTIONAL_FIELDS:
            if _has_attr(Proveedor, f):
                w = QLineEdit(dlg) if f not in ("notas","observaciones") else QTextEdit(dlg)
                inputs[f] = w
                form.addRow(f.replace("_"," ").capitalize()+":", w)
        v.addLayout(form)
        hb = QHBoxLayout(); b_ok = QPushButton("Guardar"); b_no = QPushButton("Cancelar")
        hb.addWidget(b_ok); hb.addWidget(b_no); v.addLayout(hb)
        b_ok.clicked.connect(dlg.accept); b_no.clicked.connect(dlg.reject)
        if dlg.exec_() != QDialog.Accepted: return
        nombre = txt_nombre.text().strip()
        if not nombre: QMessageBox.warning(self, "Proveedor", "Ingrese un nombre."); return
        with SessionLocal() as s:
            p = Proveedor(); p.nombre = nombre
            for k,w in inputs.items():
                val = w.toPlainText().strip() if isinstance(w, QTextEdit) else w.text().strip()
                if hasattr(p,k): setattr(p,k,val)
            s.add(p); s.commit(); pid = int(p.id)
        self._cargar_proveedores()
        for i in range(self.cbo_prov.count()):
            if int(self.cbo_prov.itemData(i)) == pid:
                self.cbo_prov.setCurrentIndex(i); break
        QMessageBox.information(self, "Proveedor", "Proveedor creado.")

    def _editar_proveedor(self):
        pid = int(self.cbo_prov.currentData() or 0)
        if not pid: QMessageBox.warning(self,"Proveedor","Seleccione un proveedor."); return
        with SessionLocal() as s:
            p = s.get(Proveedor, pid)
            if not p: QMessageBox.warning(self,"Proveedor","No encontrado."); return
        dlg = QDialog(self); dlg.setWindowTitle("Editar proveedor"); dlg.resize(520, 420)
        v = QVBoxLayout(dlg); form = QFormLayout()
        txt_nombre = QLineEdit(dlg); form.addRow("Nombre:", txt_nombre)
        inputs = {}
        for f in _PROV_OPTIONAL_FIELDS:
            if _has_attr(Proveedor, f):
                w = QLineEdit(dlg) if f not in ("notas","observaciones") else QTextEdit(dlg)
                inputs[f] = w; form.addRow(f.replace("_"," ").capitalize()+":", w)
        v.addLayout(form)
        hb = QHBoxLayout(); b_ok = QPushButton("Guardar"); b_no = QPushButton("Cancelar")
        hb.addWidget(b_ok); hb.addWidget(b_no); v.addLayout(hb)
        # precargar
        txt_nombre.setText(_get_any(p,"nombre") or "")
        for k,w in inputs.items():
            val = _get_any(p,k) or ""
            (w.setPlainText(val) if isinstance(w, QTextEdit) else w.setText(val))
        b_ok.clicked.connect(dlg.accept); b_no.clicked.connect(dlg.reject)
        if dlg.exec_() != QDialog.Accepted: return
        nombre = txt_nombre.text().strip()
        if not nombre: QMessageBox.warning(self,"Proveedor","Ingrese un nombre."); return
        with SessionLocal() as s:
            p2 = s.get(Proveedor, pid)
            p2.nombre = nombre
            for k,w in inputs.items():
                val = w.toPlainText().strip() if isinstance(w, QTextEdit) else w.text().strip()
                if hasattr(p2,k): setattr(p2,k,val)
            s.commit()
        QMessageBox.information(self, "Proveedor", "Proveedor actualizado.")

    def _ver_proveedor(self):
        pid = int(self.cbo_prov.currentData() or 0)
        if not pid: QMessageBox.information(self,"Proveedor","Seleccione un proveedor."); return
        with SessionLocal() as s:
            p = s.get(Proveedor, pid)
        info = [f"ID: {getattr(p,'id','')}", f"Nombre: {getattr(p,'nombre','')}"]
        for f in _PROV_OPTIONAL_FIELDS:
            if _has_attr(Proveedor, f): info.append(f"{f.replace('_',' ').capitalize()}: {getattr(p,f,'') or ''}")
        QMessageBox.information(self, "Proveedor", "\n".join(info))

    # ---------- listas ----------
    def _load_listas(self):
        with SessionLocal() as s:
            stmt = (select(ListaPrecioProveedor.id, ListaPrecioProveedor.proveedor_id,
                           ListaPrecioProveedor.nombre, ListaPrecioProveedor.fecha_creacion,
                           func.count(ItemListaProveedor.id).label("n"),
                           func.max(func.coalesce(getattr(ItemListaProveedor,"updated_at",None),
                                                  getattr(ListaPrecioProveedor,"fecha_creacion",None))).label("u"))
                    .join(ItemListaProveedor, ItemListaProveedor.lista_id==ListaPrecioProveedor.id, isouter=True)
                    .group_by(ListaPrecioProveedor.id)
                    .order_by(ListaPrecioProveedor.id.desc()))
            rows = s.execute(stmt).all()
            prov_cache = {}
            def pname(pid):
                if not pid: return ""
                if pid in prov_cache: return prov_cache[pid]
                p = s.get(Proveedor, pid); prov_cache[pid] = _get_any(p,"nombre") if p else ""; return prov_cache[pid]

        self.tbl.setRowCount(len(rows))
        for r,(lid,pid,nombre,fecha,n,u) in enumerate(rows):
            chk = QTableWidgetItem("")
            chk.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
            chk.setCheckState(Qt.Unchecked)
            self.tbl.setItem(r, 0, chk)

            it = QTableWidgetItem(str(lid)); it.setData(Qt.UserRole, int(lid))
            self.tbl.setItem(r, 1, it)
            self.tbl.setItem(r, 2, QTableWidgetItem(pname(pid)))
            self.tbl.setItem(r, 3, QTableWidgetItem("" if nombre is None else str(nombre)))
            self.tbl.setItem(r, 4, QTableWidgetItem("" if not fecha else (fecha.strftime("%Y-%m-%d") if hasattr(fecha,"strftime") else str(fecha))))
            self.tbl.setItem(r, 5, QTableWidgetItem(str(n or 0)))
            self.tbl.setItem(r, 6, QTableWidgetItem(u.strftime("%Y-%m-%d %H:%M") if hasattr(u,"strftime") else (str(u) if u else "")))
        self.tbl.resizeColumnsToContents()
        self.lbl_state.setText(f"Listas: {len(rows)}")

        # Asegurar ✓ único
        def _enforce(row, col):
            if col != 0: return
            it = self.tbl.item(row, 0)
            if it and it.checkState() == Qt.Checked:
                self._clear_checks_except(row)
        try: self.tbl.cellChanged.disconnect()
        except Exception: pass
        self.tbl.cellChanged.connect(_enforce)

        # Spinners con el % de la lista seleccionada (si hay)
        lid = self._lista_sel()
        if lid:
            d, v = get_lista_percents(lid)
            self.sp_desc_l.setValue(d); self.sp_iva_l.setValue(v)

    def _open_items(self):
        lid = self._lista_sel()
        if not lid:
            QMessageBox.warning(self,"Productos","Seleccione una lista."); return
        _ItemsListaDialog(lid, self).exec_()

    def _delete_lista(self):
        lid = self._lista_sel()
        if not lid: QMessageBox.warning(self,"Eliminar","Seleccione una lista."); return
        if QMessageBox.question(self,"Confirmar","¿Eliminar esta lista y sus ítems?",
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        with SessionLocal() as s:
            s.query(ItemListaProveedor).filter_by(lista_id=lid).delete()
            s.query(ListaPrecioProveedor).filter_by(id=lid).delete()
            s.commit()
        self._load_listas()
        QMessageBox.information(self,"Eliminar","Lista eliminada.")

    # ---------- guardar % ----------
    def _on_save_lista(self):
        lid = self._lista_sel()
        if not lid:
            QMessageBox.warning(self,"Lista","Seleccione una lista."); return
        d = float(self.sp_desc_l.value()); v = float(self.sp_iva_l.value())
        save_lista_percents(lid, d, v)
        # Apertura inmediata con % forzados => aplica sí o sí lo recién guardado.
        _ItemsListaDialog(lid, self, force_list_percents=(d, v)).exec_()

    def _on_save_marca(self):
        lid = self._lista_sel()
        if not lid:
            QMessageBox.warning(self,"Marca","Seleccione una lista."); return
        m = (self.txt_marca.text() or "").strip()
        if not m:
            QMessageBox.warning(self,"Marca","Indique la marca exacta."); return
        save_marca_percents(lid, m, self.sp_desc_m.value(), self.sp_iva_m.value())
        # Para ver el efecto por marca abrimos normal (lee overrides marca del storage)
        _ItemsListaDialog(lid, self).exec_()

    def _on_del_marca(self):
        lid = self._lista_sel()
        if not lid:
            QMessageBox.warning(self,"Marca","Seleccione una lista."); return
        m = (self.txt_marca.text() or "").strip()
        if not m:
            QMessageBox.warning(self,"Marca","Indique la marca exacta."); return
        delete_marca_percents(lid, m)
        _ItemsListaDialog(lid, self).exec_()

    # ---------- importar Excel ----------
    def _importar_excel(self):
        pid = int(self.cbo_prov.currentData() or 0)
        if not pid: QMessageBox.warning(self,"Proveedor","Seleccione un proveedor."); return

        path, _ = QFileDialog.getOpenFileName(self, "Seleccionar Excel", "", "Excel (*.xlsx *.xlsm *.xltx *.xltm)")
        if not path: return

        try:
            wiz = _ImportWizardDialog(path, self)
        except Exception as e:
            QMessageBox.critical(self,"Excel",f"No se pudo abrir el archivo.\n{e}"); return
        if wiz.exec_() != QDialog.Accepted: return
        sel = wiz.selections()

        try:
            wb = openpyxl.load_workbook(path, data_only=True); ws = wb[sel["sheet"]]
        except Exception as e:
            QMessageBox.critical(self,"Excel",f"No se pudo abrir la hoja.\n{e}"); return

        header = int(sel["header_row"])
        rows = list(ws.iter_rows(values_only=True))
        data_rows = []
        for i,row in enumerate(rows, start=1):
            vals = [("" if v is None else str(v).strip()) for v in row]
            if i > header and any(v != "" for v in vals): data_rows.append(vals)

        idx_codigo = sel["idx_codigo"]; idx_desc = sel["idx_desc"]; idx_marca = sel["idx_marca"]; idx_precio = sel["idx_precio"]
        idx_pres = sel["idx_pres"]; idx_extra = sel["idx_extra"]

        if idx_codigo is None or idx_desc is None or idx_precio is None:
            QMessageBox.warning(self,"Mapeo","Faltan columnas: código, descripción, precio."); return

        nombre_archivo = os.path.splitext(os.path.basename(path))[0]
        ahora = datetime.now()

        target_lid = self._lista_tildada()
        with SessionLocal() as s:
            if target_lid is None:
                row = s.execute(
                    select(ListaPrecioProveedor.id).where(
                        ListaPrecioProveedor.proveedor_id==pid,
                        ListaPrecioProveedor.nombre==nombre_archivo
                    ).limit(1)
                ).first()
                if row:
                    lid_name = int(row[0])
                    if QMessageBox.question(self,"Reemplazar",
                        f"Ya existe una lista llamada “{nombre_archivo}”. ¿Reemplazarla?",
                        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes) == QMessageBox.Yes:
                        target_lid = lid_name

            if target_lid is None:
                lst = ListaPrecioProveedor(proveedor_id=pid, nombre=nombre_archivo, fecha_creacion=ahora)
                s.add(lst); s.flush()
                lid = int(lst.id)
                keep_id = False
            else:
                lst = s.get(ListaPrecioProveedor, int(target_lid))
                if not lst:
                    QMessageBox.warning(self,"Reemplazar","La lista tildada no existe."); return
                s.query(ItemListaProveedor).filter_by(lista_id=lst.id).delete()
                lst.nombre = nombre_archivo
                lst.fecha_creacion = ahora
                lid = int(lst.id)
                keep_id = True

            code_attr  = "producto_codigo" if hasattr(ItemListaProveedor,"producto_codigo") else "codigo"
            has_prec   = hasattr(ItemListaProveedor,"prec")
            has_precio = hasattr(ItemListaProveedor,"precio")
            rubro_attr = "rubro_detectado" if hasattr(ItemListaProveedor,"rubro_detectado") else None

            insertados = 0
            for r in data_rows:
                codigo = (r[idx_codigo] or "").strip() if idx_codigo is not None and idx_codigo < len(r) else ""
                desc   = (r[idx_desc]   or "").strip() if idx_desc   is not None and idx_desc   < len(r) else ""
                marca  = (r[idx_marca]  or "").strip() if idx_marca  is not None and idx_marca  < len(r) else ""
                precio = (r[idx_precio] or "").strip() if idx_precio is not None and idx_precio < len(r) else ""
                pres   = (r[idx_pres]   or "").strip() if idx_pres   is not None and idx_pres   < len(r) else ""
                extra  = (r[idx_extra]  or "").strip() if idx_extra  is not None and idx_extra  < len(r) else ""

                if desc == "" and (codigo == "" or precio == ""): continue

                if not codigo and desc:
                    codigo = desc

                item = ItemListaProveedor()
                setattr(item, "lista_id", lid)
                setattr(item, code_attr, codigo)
                setattr(item, "descripcion", desc)
                setattr(item, "marca", marca)
                if hasattr(item, "presentacion"): setattr(item, "presentacion", pres)
                if hasattr(item, "info_extra"): setattr(item, "info_extra", extra)

                val = _num(precio)
                if has_prec:   setattr(item, "prec", float(val))
                if has_precio: setattr(item, "precio", float(val))
                if rubro_attr: setattr(item, rubro_attr, None)
                s.add(item); insertados += 1
            s.commit()

        self._load_listas()
        QMessageBox.information(self, "Importación",
            f"{'Lista reemplazada' if keep_id else 'Nueva lista creada'} (ID {lid}). Ítems: {insertados}")

def _num(x):
    s = str(x).strip()
    if s == "": return 0.0
    try:
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        try: return float(x)
        except Exception: return 0.0
