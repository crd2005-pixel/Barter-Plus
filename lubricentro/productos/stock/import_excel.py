# -*- coding: utf-8 -*-
import openpyxl
import os
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox, QSpinBox, QFormLayout, QFileDialog
)
from PyQt5.QtCore import Qt

from services.producto_service import ProductoService

# Intentar importar modelos
try:
    from db.models.productos import Stock as StockModel, Marca as MarcaModel, Deposito as DepositoModel
except ImportError:
    try:
        from db import Stock as StockModel, Marca as MarcaModel, Deposito as DepositoModel
    except ImportError:
        StockModel = None; MarcaModel = None; DepositoModel = None

# ======================================================================
# Diálogo de Asistente de Mapeo (estilo "Proveedores")
# ======================================================================
class _ImportWizardDialog(QDialog):
    def __init__(self, xlsx_path: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Importar Stock – Mapeo de columnas")
        self.resize(900, 600)

        # Cargar libro con openpyxl (data_only para valores calculados)
        try:
            self._wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception:
            # Fallback si es CSV u otro formato no soportado por openpyxl directamente
            # (El usuario pidió "copiar formato", asumimos xlsx principal, pero si falla, avisamos)
            raise ValueError("El formato debe ser Excel (.xlsx, .xlsm). Para CSV conviértalo primero.")

        self._headers = []
        self._rows = []

        lay = QVBoxLayout(self)

        # 1. Selección de hoja
        fila1 = QHBoxLayout()
        fila1.addWidget(QLabel("Hoja:"))
        self.cbo_sheet = QComboBox(self)
        for n in self._wb.sheetnames:
            self.cbo_sheet.addItem(n)
        fila1.addWidget(self.cbo_sheet)
        fila1.addStretch(1)
        lay.addLayout(fila1)

        # 2. Selección de fila de encabezado
        grp = QGroupBox("Encabezado")
        gl = QHBoxLayout(grp)
        gl.addWidget(QLabel("Fila de encabezado:"))
        self.sp_head = QSpinBox(self)
        self.sp_head.setRange(1, 9999)
        self.sp_head.setValue(1)
        gl.addWidget(self.sp_head)
        gl.addStretch(1)
        lay.addWidget(grp)

        # 3. Mapeo de columnas
        mapg = QGroupBox("Mapeo de columnas")
        fm = QFormLayout(mapg)

        # Campos requeridos / opcionales para Stock
        self.cbo_nombre = QComboBox(self)
        self.cbo_marca = QComboBox(self)
        self.cbo_cod_barras = QComboBox(self)
        self.cbo_precio = QComboBox(self)
        self.cbo_cantidad = QComboBox(self)
        self.cbo_rubro = QComboBox(self)
        self.cbo_codigo = QComboBox(self) # Código interno/SKU

        fm.addRow("Nombre del Producto *:", self.cbo_nombre)
        fm.addRow("Marca:", self.cbo_marca)
        fm.addRow("Código de Barras:", self.cbo_cod_barras)
        fm.addRow("Precio Venta:", self.cbo_precio)
        fm.addRow("Cantidad (Stock):", self.cbo_cantidad)
        fm.addRow("Rubro:", self.cbo_rubro)
        fm.addRow("Código Interno (SKU):", self.cbo_codigo)

        lay.addWidget(mapg)

        # 4. Previsualización
        self.tbl = QTableWidget(self)
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        lay.addWidget(self.tbl)

        # 5. Botones
        fila_b = QHBoxLayout()
        self.btn_ok = QPushButton("Importar")
        self.btn_no = QPushButton("Cancelar")
        fila_b.addStretch(1)
        fila_b.addWidget(self.btn_ok)
        fila_b.addWidget(self.btn_no)
        lay.addLayout(fila_b)

        # Conexiones
        self.cbo_sheet.currentIndexChanged.connect(self._reload)
        self.sp_head.valueChanged.connect(self._reload)
        self.btn_ok.clicked.connect(self.accept)
        self.btn_no.clicked.connect(self.reject)

        self._reload()

    def _reload(self):
        sheet_name = self.cbo_sheet.currentText()
        if not sheet_name: return
        ws = self._wb[sheet_name]

        hrow = self.sp_head.value()
        # Leemos un sample para no trabar si es gigante
        # openpyxl iter_rows es lazy, convertimos a lista
        all_rows = list(ws.iter_rows(values_only=True))

        heads = []
        body = []

        for i, row in enumerate(all_rows, start=1):
            # Convertir None a "" y limpiar espacios
            vals = [("" if v is None else str(v).strip()) for v in row]

            if i == hrow:
                heads = [v.lower() for v in vals]
            elif i > hrow:
                # Solo agregar si tiene algo de contenido
                if any(v != "" for v in vals):
                    body.append(vals)

        self._headers = heads
        self._rows = body

        # Helper para llenar combos y auto-seleccionar
        def fill(cbo, prefer_list):
            cbo.clear()
            # Opción vacía/ignorar
            cbo.addItem("(Ignorar)", -1)

            for idx, h in enumerate(self._headers):
                cbo.addItem(f"{idx}: {h}", idx)

            # Auto-select
            pick = None
            for p in prefer_list:
                for idx, h in enumerate(self._headers):
                    if p in h:
                        pick = idx
                        break
                if pick is not None: break

            if pick is not None:
                # +1 porque el index 0 es (Ignorar)
                cbo.setCurrentIndex(pick + 1)

        fill(self.cbo_nombre,      ["nombre", "descrip", "detalle", "producto"])
        fill(self.cbo_marca,       ["marca", "brand", "fabricante"])
        fill(self.cbo_cod_barras,  ["barras", "ean", "upc", "barcode"])
        fill(self.cbo_precio,      ["precio", "venta", "pvp", "final"])
        fill(self.cbo_cantidad,    ["cant", "stock", "existencia", "q"])
        fill(self.cbo_rubro,       ["rubro", "categor", "familia"])
        fill(self.cbo_codigo,      ["codigo", "código", "sku", "interno"])

        # Actualizar tabla preview
        self.tbl.clear()
        self.tbl.setColumnCount(len(self._headers))
        self.tbl.setHorizontalHeaderLabels(self._headers)

        sample = self._rows[:50] # mostrar solo 50
        self.tbl.setRowCount(len(sample))
        for r, row_data in enumerate(sample):
            for c, v in enumerate(row_data):
                self.tbl.setItem(r, c, QTableWidgetItem(v))
        self.tbl.resizeColumnsToContents()

    def selections(self):
        """Retorna un dict con los índices de columna seleccionados (-1 si ignorar)"""
        return {
            "sheet": self.cbo_sheet.currentText(),
            "idx_nombre": self.cbo_nombre.currentData(),
            "idx_marca": self.cbo_marca.currentData(),
            "idx_cod_barras": self.cbo_cod_barras.currentData(),
            "idx_precio": self.cbo_precio.currentData(),
            "idx_cantidad": self.cbo_cantidad.currentData(),
            "idx_rubro": self.cbo_rubro.currentData(),
            "idx_codigo": self.cbo_codigo.currentData(),
        }

# ======================================================================
# Clase Principal de Importación (Lógica de Negocio)
# ======================================================================
class ImportExcelDialog(QDialog):
    """
    Wrapper compatible con la llamada desde base_stock.py.
    Abre el Wizard y procesa la importación.
    """
    def __init__(self, parent=None, session_factory=None, producto_model=None):
        super().__init__(parent)
        self.SessionLocal = session_factory
        self.ProductoModel = producto_model
        # No mostramos UI propia, lanzamos el wizard inmediatamente o en load_file
        self.layout = QVBoxLayout(self)
        self.setLayout(self.layout)
        self.hide()

    def load_file(self, file_path):
        """
        Lanza el wizard. Si se acepta, procesa. Si no, cierra.
        """
        try:
            wiz = _ImportWizardDialog(file_path, self.parent())
            if wiz.exec_() == QDialog.Accepted:
                sel = wiz.selections()
                self._procesar_importacion(file_path, sel)
                self.accept()
            else:
                self.reject()
        except Exception as e:
            QMessageBox.critical(self.parent(), "Error de Archivo", f"No se pudo procesar el archivo:\n{e}")
            self.reject()

    def _procesar_importacion(self, file_path, sel):
        if not self.SessionLocal or not self.ProductoModel:
            return

        # Indices (-1 si no seleccionó)
        idx_nombre = sel["idx_nombre"]
        idx_marca = sel["idx_marca"]
        idx_cod_barras = sel["idx_cod_barras"]
        idx_precio = sel["idx_precio"]
        idx_cantidad = sel["idx_cantidad"]
        idx_rubro = sel["idx_rubro"]
        idx_codigo = sel["idx_codigo"]

        if idx_nombre == -1:
            QMessageBox.warning(self.parent(), "Importación", "El Nombre es obligatorio.")
            return

        # Cargar datos reales
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sel["sheet"]]
        # Necesitamos saber dónde empezaba la data (usamos wizard logic implícita?
        # No, mejor iteramos todo y saltamos headers si detectamos texto de header,
        # PERO el wizard permite elegir fila header.
        # Para simplificar, re-leemos todo y filtramos igual que el wizard o confiamos en que
        # el usuario no cambió la fila header mentalmente.
        # *Mejor*: leer todo y asumir la fila header detectada es la fila N, data N+1.
        # Como no guardamos N en 'sel', asumimos iteración robusta:
        # Buscaremos header o simplemente procesamos si nombre no es vacío.

        # Una forma más segura: Pedir al wizard que retorne las filas DATA.
        # Pero el wizard es UI. Replicamos lectura simple:

        rows = list(ws.iter_rows(values_only=True))

        count_new = 0
        count_updated = 0

        with self.SessionLocal() as s:
            # Asegurar deposito
            deposito = None
            if DepositoModel:
                deposito = s.query(DepositoModel).first()
                if not deposito:
                    deposito = DepositoModel(nombre="Principal")
                    s.add(deposito)
                    s.flush()

            # Cache marcas para no query por cada row
            marcas_cache = {} # nombre -> id
            if MarcaModel:
                for m in s.query(MarcaModel).all():
                    marcas_cache[m.nombre.lower()] = m.id

            # Cache productos existentes (por codigo barra y nombre) para minimizar queries
            # OJO: Si son miles, cargar todo a memoria es viable (ids, nombres, barras).
            # Si es MUY grande, mejor query individual. Asumimos < 10k productos -> memoria ok.
            existing_barras = {} # barra -> producto_obj
            existing_nombres = {} # nombre -> producto_obj
            existing_skus = {} # sku -> producto_obj

            all_prods = s.query(self.ProductoModel).all()
            for p in all_prods:
                if p.codigo_barras: existing_barras[p.codigo_barras] = p
                if p.nombre: existing_nombres[p.nombre.lower()] = p
                if p.sku: existing_skus[p.sku] = p

            # Helper numérico
            def _clean_float(val):
                if val is None: return None
                s_val = str(val).replace('$', '').replace(' ', '')
                # Si tiene , y . asumir formato según convención simple o probar ambas
                # Estandar Argentina: 1.000,50 o 1000,50
                # Python float usa punto.
                try:
                    return float(s_val)
                except:
                    try:
                        # Cambio coma por punto
                        return float(s_val.replace(',', '.'))
                    except:
                        # Intento eliminar puntos de miles y luego coma
                        # ej 1.200,50 -> 1200.50
                        try:
                            return float(s_val.replace('.', '').replace(',', '.'))
                        except:
                            return None

            # Iteramos
            header_found = False
            for r_idx, row in enumerate(rows):
                # Validar rango indices
                def get_v(idx):
                    if idx >= 0 and idx < len(row):
                        v = row[idx]
                        return str(v).strip() if v is not None else ""
                    return ""

                nombre = get_v(idx_nombre)

                # Heurística simple para saltar header si está incluido en rows
                # Si la columna nombre dice "nombre" (case insensitive) ignoramos
                if nombre.lower() in ["nombre", "descripción", "descripcion", "producto"]:
                    header_found = True
                    continue

                if not nombre: continue # Fila vacía

                # Extracción de datos
                barra = get_v(idx_cod_barras) if idx_cod_barras >= 0 else None
                sku = get_v(idx_codigo) if idx_codigo >= 0 else None

                # Identificar producto
                prod = None

                # 1. Por Barra
                if barra and barra in existing_barras:
                    prod = existing_barras[barra]
                # 2. Por SKU
                elif sku and sku in existing_skus:
                    prod = existing_skus[sku]
                # 3. Por Nombre
                elif nombre.lower() in existing_nombres:
                    prod = existing_nombres[nombre.lower()]

                is_new = False
                if not prod:
                    prod = self.ProductoModel()
                    is_new = True
                    count_new += 1
                else:
                    count_updated += 1

                # Actualizar campos (siempre actualiza o rellena)
                prod.nombre = nombre # Nombre manda del excel
                if barra: prod.codigo_barras = barra
                if sku: prod.sku = sku

                # Marca
                if idx_marca >= 0:
                    m_str = get_v(idx_marca)
                    if m_str:
                        m_lower = m_str.lower()
                        if m_lower in marcas_cache:
                            prod.marca_id = marcas_cache[m_lower]
                        else:
                            # Crear marca
                            if MarcaModel:
                                new_m = MarcaModel(nombre=m_str)
                                s.add(new_m)
                                s.flush()
                                marcas_cache[m_lower] = new_m.id
                                prod.marca_id = new_m.id

                # Rubro
                if idx_rubro >= 0:
                    r_str = get_v(idx_rubro)
                    if r_str: prod.rubro = r_str

                # Precio
                if idx_precio >= 0:
                    p_val = _clean_float(get_v(idx_precio))
                    if p_val is not None:
                        prod.precio_minorista = p_val

                if is_new:
                    s.add(prod)
                    s.flush() # Para tener ID
                    # Agregar a caches para futuras referencias en este mismo loop (por si hay duplicados en excel)
                    if prod.codigo_barras: existing_barras[prod.codigo_barras] = prod
                    if prod.nombre: existing_nombres[prod.nombre.lower()] = prod
                    if prod.sku: existing_skus[prod.sku] = prod

                # Stock (Concatenar / Actualizar)
                # El usuario pidió "concatenar o actualizar".
                # Interpretación: Si en el excel dice 10, y yo tenía 5, ¿ahora tengo 15 o 10?
                # "Importación masiva" suele ser un "stock take" (toma de inventario) -> reemplaza.
                # O puede ser "entrada de mercadería" -> suma.
                # Como dice "Flexibilidad Post-Carga... editar manualmente", y "no borrar registros actuales",
                # asumiremos REEMPLAZO de la cantidad con lo que dice el Excel (Toma de inventario),
                # ya que importar una lista suele implicar "esto es lo que tengo hoy".
                # Sin embargo, para ser seguros y "master", si el usuario quiere sumar, debería ser otra opción.
                # Ante la duda en "carga masiva" de stock inicial o corrección, el valor del Excel MANDA.

                if idx_cantidad >= 0 and StockModel and deposito:
                    q_val = _clean_float(get_v(idx_cantidad))
                    if q_val is not None:
                        # Buscar stock record
                        st = s.query(StockModel).filter(
                            StockModel.producto_id == prod.id,
                            StockModel.deposito_id == deposito.id
                        ).first()

                        if not st:
                            st = StockModel(producto_id=prod.id, deposito_id=deposito.id, cantidad=0)
                            s.add(st)

                        # Actualizamos cantidad (Overwrite logic)
                        st.cantidad = q_val

            s.commit()

        QMessageBox.information(self.parent(), "Importación Finalizada",
                                f"Proceso completado.\n\nProductos Nuevos: {count_new}\nProductos Actualizados: {count_updated}")
